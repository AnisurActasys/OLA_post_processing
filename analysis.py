from pathlib import Path
from openpyxl import Workbook, load_workbook
from pyparsing import col
import matplotlib.pyplot as plt
import numpy as np
from statistics import mean, stdev
import xlwt
from xlwt import Workbook

#Locate and load Excel file
excel_path = Path.cwd() / "cleaning_effectiveness_results.xlsx"
wb = load_workbook(excel_path, data_only=True)
ws = wb.active
max_row = ws.max_row
max_col = ws.max_column

#All data from Excel file will be captured in one dictionary variable: cleaning_data{}
cleaning_data = {}
derivative_dict = {}
for j in range(1, max_col, 3):
    table_values = {}
    derivative_vals = []
    parameter = ws.cell(row = 1, column = j).value
    table_values[ws.cell(row = 3, column = j).value] = [0, 0, 0]
    i = 4
    while(True):
        if ws.cell(row = i, column = j).value == None:
            break
        else:
            derivative = (ws.cell(row = i, column = j+1).value - ws.cell(row = i-1, column = j+1).value) / ((ws.cell(row = i, column = j).value - ws.cell(row = i-1, column = j).value))
            derivative_vals.append(derivative)
            table_values[ws.cell(row = i, column = j).value] = [ws.cell(row = i, column = j+1).value, ws.cell(row = i, column = j+2).value, derivative]
            i += 1
    cleaning_data[parameter] = table_values
    derivative_dict[parameter] = derivative_vals

#Seperate out the different operating conditions being compared
cases = list(cleaning_data.keys())

#Write out the data to a different Excel File. This time each case will get its own sheet 
wb = Workbook() 
for case in cases:
    curr_sheet = wb.add_sheet(case)
    times = list(cleaning_data[case].keys())
    curr_sheet.write(0, 0, case)
    curr_sheet.write(1, 0, "Time (s)")
    curr_sheet.write(1, 1, "Cleaning Effectiveness (%)")
    curr_sheet.write(1, 2, "Energy Consumed (kJ)")
    curr_sheet.write(1, 3, "Cleaning Rate Speed (%/s)")
    i = 2
    for time in times:
        clean_eff = cleaning_data[case][time][0]
        energy_cons = cleaning_data[case][time][1]
        clean_derivative = cleaning_data[case][time][2]
        
        curr_sheet.write(i, 0, round(time, 2))
        curr_sheet.write(i, 1, round(clean_eff, 2))
        curr_sheet.write(i, 2, round(energy_cons, 2))
        curr_sheet.write(i, 3, round(clean_derivative, 2))
        i += 1

wb.save('seperated_cleaning_effectiveness_results.xls') #Saving as xlsx file was causing errors so xls is used for now

#80% cleaning effectiveness will be the cutoff for the analysis
#Create a new dictionary that only holds values from when actuator until the cutoff point
to_80 = {}
energy_dict = {}
inverse_energy_dict = {}
time_dict = {}
clean_speed_dict = {}
cutoff = 80

#The actuator is assumed to turn on when the derivative value for cleaning is at its maximum
#Optimizing speed is a maximizing operation while opimitizing energy consumption is a minimizing operation
#In order to create a "performance score", the inverse of the energy will be used to both values can be maximized
for case in cases:
    time_keys = list(cleaning_data[case].keys())
    derivatives = list(derivative_dict[case])
    time_and_energy = {}
    for j in range(len(time_keys)):
        time_key = time_keys[j]
        clean_percent = cleaning_data[case][time_key][0]
        energy = cleaning_data[case][time_key][1]
        derivative = cleaning_data[case][time_key][2]
        time_to_80 = None
        energy_to_80 = None
        if derivative == max(derivatives):
            reference_time = time_key
        if clean_percent == cutoff:
            time_to_80 = time_key - reference_time #By subtracting the reference time, the cleaning duration is calculated
            energy_to_80 = (energy / time_key) * time_to_80 / 1000
            break
        elif clean_percent > cutoff: #Linear interpolation
            time_to_80 = (time_key - time_keys[j-1]) / (clean_percent - cleaning_data[case][time_keys[j-1]][0]) * (cutoff - cleaning_data[case][time_keys[j-1]][0]) + time_keys[j-1] - reference_time
            energy_to_80 = energy / time_key * time_to_80 / 1000
            break
    if time_to_80 != None: #Don't include cases that don't reach cutoff value. Will cause division by zero error
        time_and_energy["Time"] = time_to_80 
        time_and_energy["Energy"] = energy_to_80
        to_80[case] = time_and_energy
        energy_dict[case] = energy_to_80
        inverse_energy_dict[case] = 1 / energy_to_80 
        time_dict[case] = time_to_80
        clean_speed_dict[case] = cutoff / time_to_80

#Standard score standardization shifts the mean to 0 and ensures that the stdev is 1.0. 
def standard_score_normalization(dictionary): 
    """Take a dataset in the form of a dicitonary as an arg and return a standardized dataset."""
    cases = list(dictionary.keys())
    new_dict = {}
    for case in cases:
        mean_value = mean(list(dictionary.values()))
        val = dictionary[case]
        sdev_val = stdev(list(dictionary.values()))

        normalized_val = (val-mean_value) / sdev_val
        new_dict[case] = normalized_val

    return new_dict

#min-max scaling method brings values to the range of [0,1]. 
def min_max_normalization(dictionary):
    """Take a dataset in the form of a dicitonary as an arg and return a normalized dataset."""
    cases = list(dictionary.keys())
    new_dict = {}
    for case in cases:
        min_value = min(list(dictionary.values()))
        max_value = max(list(dictionary.values()))
        val = dictionary[case]

        normalized_val = (val-min_value) / (max_value - min_value)
        new_dict[case] = normalized_val

    return new_dict


normalized_inverse_energies_dict = min_max_normalization(inverse_energy_dict)
normalized_cleaning_speeds_dict = min_max_normalization(clean_speed_dict)

cases = list(energy_dict.keys())
energies = list(energy_dict.values())
inverse_energies = list(inverse_energy_dict.values())
times = list(time_dict.values())
cleaning_speeds = list(clean_speed_dict.values())

normalized_inverse_energies = list(normalized_inverse_energies_dict.values())
normalized_cleaning_speeds = list(normalized_cleaning_speeds_dict.values())


def OLA_plots(cases, var1, var2, axis_labels, legend_labels, title, figure_number):
    """Create a bar plot for data sets plotted on the same x-axis with different y-axes."""
    fig, ax = plt.subplots(figsize = (16,6))
    labels = cases
    x = np.arange(len(cases))
    ax2 = ax.twinx()

    ax.set_xlabel("Input Parameters")
    ax.set_ylabel(axis_labels[0])
    ax2.set_ylabel(axis_labels[1])

    color = ['red', 'royalblue']
    width = 0.25

    p1 = ax.bar(x-width, var1, width = width, color = color[0], align = 'edge', label = legend_labels[0])
    p2 = ax2.bar(x, var2, width = width, color = color[1], align = 'edge', label = legend_labels[1])

    lns = [p1, p2]
    ax.legend(handles = lns, loc = 'best')
    ax.set_xticks(x)
    ax.set_xticklabels(labels)
    ax.set_title(title,fontsize=18, weight='bold')

    return plt.figure(figure_number)


#user_input = input("Would you like to see results graph? Enter 'y' for yes or 'n' for no: ")
#if user_input == 'y'
#    plt.show()

speed_weight = 0.6
energy_weight = 0.4
weight_array = [speed_weight, energy_weight]

rankings_dict = {}


def get_scores(dictionary_array, weight_array):
    """Take in an array of dictionaries containing the different scoring criteria as the first argument and an array of their respective weights as the second argument. Return a score for each operating case. Dictionaries must have the same keys."""
    scores_dict = {}
    inv_scores_dict = {}
    for case in cases:
        val_array = []
        for dictionary in dictionary_array:
            val_array.append(dictionary[case])
        array_sum = 0
        for i in range(len(weight_array)):
            array_sum += val_array[i] * weight_array[i]
        scores_dict[case] = array_sum
        inv_scores_dict[array_sum] = case #flip the key-value relationship so a sorted dictionary can be created later
    return [scores_dict, inv_scores_dict]


scores = get_scores([normalized_cleaning_speeds_dict, normalized_inverse_energies_dict], weight_array)
scores_dict = scores[0]
inv_scores_dict = scores[1]


sorted_ranks = sorted(inv_scores_dict, reverse=True)
sorted_cases = []
for rank in sorted_ranks:
    sorted_cases.append(inv_scores_dict[rank])



#p1 = OLA_plots(cases, cleaning_speeds, times, ["Clean Speed (%/s)", "Cleaning Time (s)"], ["Clean Speed", "Time"], "Cleaning Speeds and Cleaning imes", 1)
#p2 = OLA_plots(cases, energies, times, ["Energy Consumed (kJ)", "Cleaning Time (s)"], ["Energy", "Time"], "Energy Consumption and Cleaning Times", 2)
#p3 = OLA_plots(cases, energies, cleaning_speeds, ["Energy Consumed (kJ)", "Cleaning Speed (%/s)"], ["Energy", "Clean Speed"], "Energy Consumption and leaning Speeds", 3)
#p4 = OLA_plots(cases, normalized_inverse_energies, normalized_cleaning_speeds, ["Normalized Inverse of Energy Consumed", "Normalized Cleaning Speed"], ["Normalized Inverse Energy", "#Normalized Cleaning Speed"], "Normalized Inverse of Energy Consumption and Cleaning Speed", 4)#
#fig = plt.figure(figsize= (16,6))
#plt.bar(sorted_cases, sorted_ranks, color = 'red', width = 0.25)#
#plt.xlabel("Input Parameters")
#plt.ylabel("Score")
#plt.title("Scores for Different Operating Conditions")
#plt.show()


#def post_process(data_out_edge, ip):
#    derivative_array = [0]
#    time_array = list(data_out_edge['time_array'])
#    cleaning_eff_array = list(data_out_edge['CleaningEffNorm1'])
#    energy_array = list(data_out_edge['Energy'])
#    for i in range(1, len(time_array)):
#        derivative_array.append((cleaning_eff_array[i] - cleaning_eff_array[i-1]) / (time_array[i] - time_array[i-1]))
#    data_out_edge['cleaning_speeds'] = derivative_array
#    max_speed = max(derivative_array)
#
#    for i in range(time_array):
#        time = time_array[i]
#        speed = derivative_array[i]
#        clean_eff = cleaning_eff_array[i]
#        cutoff = ip['cutoff_value']
#        if speed == speed:
#            ref_time = time[i]
#            ref_energy = energy_array[i]
#        if clean_eff >= cutoff:
#            prev_clean = cleaning_eff_array[i-1]
#            prev_time = time_array[i-1]
#            cutoff_time = (time - prev_time) / (clean_eff - prev_clean) * (cutoff - prev_clean) + prev_time
#            energy_consumption = ref_energy/ref_time * cutoff_time
#










